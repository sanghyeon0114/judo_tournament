import pandas as pd
import random
import shutil
import os
from openpyxl import load_workbook, Workbook

def getName(df, index):
    print(df.loc[index]['name'])

def getTeam(df, index):
    print(df.loc[index]['team'])

def create_round_one(df):
    team_list = list(df['team'].unique())
    index_list = []
    for i in range(len(team_list)):
        index_list.append([])
        
    user_dict = dict(zip(team_list, index_list))
    
    for index, (name, team) in df.iterrows():
        user_dict[team].append(index)
    
    tournament = []
    while len(team_list) != 0:
        team_len = len(team_list)
        team1 = max(user_dict, key=lambda team: len(user_dict[team]))
        team1_user_list = user_dict[team1]
    
        user1 = random.choice(team1_user_list)
        team1_user_list.remove(user1)

        if len(team1_user_list) == 0:
            user_dict.pop(team1)
            team_list.remove(team1)

        team_len = len(team_list)
        if team_len == 0:
            tournament.append((user1, None))
            break
        
        team2 = team_list[random.randrange(team_len)]

        if team_len != 1:
            while team1 == team2:
                team2 = team_list[random.randrange(team_len)]
        team2_user_list = user_dict[team2]

        user2 = random.choice(team2_user_list)
        team2_user_list.remove(user2)

        if len(team2_user_list) == 0:
            user_dict.pop(team2)
            team_list.remove(team2)

        tournament.append((user1, user2))

    random.shuffle(tournament)
    return tournament

def modifyWorkbook(ws, teamCol, nameCol, team, name):
    ws[teamCol].value = team
    ws[nameCol].value = name

def modifyOneRoundInWorkbook(df, ws, round, excelCol1, excelCol2):
    user1 = df.loc[round[0]]
    user2 = df.loc[round[1]]
    modifyWorkbook(ws, excelCol1[0], excelCol1[1], user1['team'], user1['name'])
    modifyWorkbook(ws, excelCol2[0], excelCol2[1], user2['team'], user2['name'])

def modifyOnePeopleInWorkbook(df, ws, round, excelCol):
    user = df.loc[round[0]]
    modifyWorkbook(ws, excelCol[0], excelCol[1], user['team'], user['name'])

def divideOneRound(df, tournament_list):
    new_tournament_list = tournament_list
    for round in new_tournament_list:
        if df.loc[round[0]]['team'] == df.loc[round[1]]['team']:
            new_tournament_list.remove(round)
            new_tournament_list.append((round[0], None))
            new_tournament_list.append((round[1], None))
            return tournament_list
    tmp_round = new_tournament_list[0]
    new_tournament_list.remove(new_tournament_list[0])
    new_tournament_list.append((tmp_round[0], None))
    new_tournament_list.append((tmp_round[1], None))
    return tournament_list

file_name = 'list.xlsx'
df = pd.read_excel(file_name)
df.columns = ["name", "team"]

user_len = len(df)
tournament_result = create_round_one(df)

if user_len < 4:
    print("인원이 부족합니다. ( 최소 인원 : 4 )")
    quit()

file_name = input("파일 이름 입력 : ")

path = "./outputs"
file_list = os.listdir(path)

while file_name+".xlsx" in file_list:
    print("동일한 파일명이 있습니다. 다른 파일명으로 바꿔주세요.")
    file_name = input("파일 이름 입력 : ")

from_file_path = './files/tou%s.xlsx' % (user_len)
to_file_path = './outputs/tmp_excel_file.xlsx'
shutil.copyfile(from_file_path, to_file_path) 

 
wb = load_workbook(to_file_path)
ws = wb['Sheet1']

title_name = input("타이틀 입력 : ")
tmp_title = ws['D5'].value
tmp_title = tmp_title.replace("title", title_name)
ws['D5'].value = tmp_title

if user_len == 8:
    first_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, first_round, ('B14', 'D14'), ('B18', 'D18'))

    second_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, second_round, ('B22', 'D22'), ('B26', 'D26'))

    third_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, third_round, ('Q14', 'P14'), ('Q18', 'P18'))

    forth_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, forth_round, ('Q22', 'P22'), ('Q26', 'P26'))

elif user_len == 7:    
    for one_round in tournament_result:
        if None in one_round:
            tournament_result.remove(one_round)
            modifyOnePeopleInWorkbook(df, ws, one_round, ('B24', 'D24'))
            break

    first_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, first_round, ('B14', 'D14'), ('B18', 'D18'))

    third_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, third_round, ('Q14', 'P14'), ('Q18', 'P18'))

    forth_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, forth_round, ('Q22', 'P22'), ('Q26', 'P26'))

elif user_len == 6:
    tournament_result = divideOneRound(df, tournament_result)
    print(tournament_result)

    for one_round in tournament_result:
        if None in one_round:
            tournament_result.remove(one_round)
            modifyOnePeopleInWorkbook(df, ws, one_round, ('B22', 'D22'))
            break
    
    for one_round in tournament_result:
        if None in one_round:
            tournament_result.remove(one_round)
            modifyOnePeopleInWorkbook(df, ws, one_round, ('Q22', 'P22'))
            break

    first_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, first_round, ('B14', 'D14'), ('B18', 'D18'))

    third_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, third_round, ('Q14', 'P14'), ('Q18', 'P18'))

elif user_len == 5:
    for one_round in tournament_result:
        if None in one_round:
            tournament_result.remove(one_round)
            modifyOnePeopleInWorkbook(df, ws, one_round, ('B22', 'D22'))
            break
        
    first_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, first_round, ('B14', 'D14'), ('B18', 'D18'))

    third_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, third_round, ('P16', 'O16'), ('P22', 'O22'))


elif user_len == 4:
    first_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, first_round, ('B14', 'D14'), ('B18', 'D18'))

    second_round = tournament_result.pop()
    modifyOneRoundInWorkbook(df, ws, second_round, ('M14', 'L14'), ('M18', 'L18'))

wb.save(to_file_path)

new_file_name = './outputs/%s.xlsx' % (file_name)
os.rename(to_file_path, new_file_name)